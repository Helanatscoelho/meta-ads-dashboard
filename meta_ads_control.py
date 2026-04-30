#!/usr/bin/env python3
"""
Meta Ads Financial Control System
Generates weekly Excel report: balance projections, campaign spend, executive summary.
Run: python meta_ads_control.py
Schedule: python scheduler.py  (every Monday 08:00)
"""

import os
import sys
import json
import time
import logging
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Tuple

try:
    import requests
except ImportError:
    print("ERROR: pip install -r requirements.txt")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install -r requirements.txt")
    sys.exit(1)

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ─── Configuration ─────────────────────────────────────────────────────────────

ACCESS_TOKEN = os.environ.get("META_ACCESS_TOKEN", "")
BASE_URL = "https://graph.facebook.com/v19.0"
REPORT_DAYS = 7
CRITICAL_DAYS = 3
WARNING_DAYS = 7
CTR_BENCHMARK = 1.0   # %
CPM_MAX = 30.0        # R$
CPC_MAX = 5.0         # R$
MAX_RETRIES = 3

ACCOUNTS = [
    {"id": "act_101403706860167",  "name": "Perfiletto"},
    {"id": "act_267770730525130",  "name": "Helena Coelho"},
    {"id": "act_2410626089217011", "name": "Conta 2410"},
    {"id": "act_1182499605271369", "name": "MM Protege"},
    {"id": "act_655917871724513",  "name": "Ergy Diesel"},
    {"id": "act_196598645524416",  "name": "Marcelo Costa"},
    {"id": "act_297842908343350",  "name": "Centro do Sorriso Arapongas"},
    {"id": "act_404664394554450",  "name": "Saniteck"},
    {"id": "act_558494692916247",  "name": "Lady Livretos"},
    {"id": "act_839314777805875",  "name": "Centro do Sorriso Campo Mourão"},
    {"id": "act_1047813853024474", "name": "Cheflera"},
    {"id": "act_1559065974904383", "name": "Clínica Moliah"},
    {"id": "act_985077316316981",  "name": "Arezzo"},
    {"id": "act_1463630264317907", "name": "Ana Capri"},
    {"id": "act_1648225505751115", "name": "Comunidade Trindade Santa"},
    {"id": "act_864538029208474",  "name": "Monique"},
    {"id": "act_557811123855687",  "name": "Loja Trindade Santa"},
    {"id": "act_2029226410831026", "name": "CA - Helena"},
    {"id": "act_1205207067848409", "name": "Centro do Sorriso Rolândia"},
    {"id": "act_1303967340832416", "name": "Odonto Bless Londrina"},
    {"id": "act_668541309479669",  "name": "Castro e Figueira"},
    {"id": "act_1908379856581665", "name": "Carol Meirelles"},
    {"id": "act_1575227266844090", "name": "Vedoi"},
    {"id": "act_877477095207362",  "name": "Star Veículos"},
]

# ─── Excel Styles ───────────────────────────────────────────────────────────────

BLUE_FILL   = PatternFill(start_color="1877F2", end_color="1877F2", fill_type="solid")
RED_FILL    = PatternFill(start_color="E53935", end_color="E53935", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
GREEN_FILL  = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
ALT_FILL    = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
WARN_LIGHT  = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
CRIT_LIGHT  = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

WHITE_BOLD  = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14, color="1877F2")

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ─── API Helpers ───────────────────────────────────────────────────────────────

def api_get(url: str, params: dict) -> dict:
    for attempt in range(MAX_RETRIES):
        try:
            r = requests.get(url, params=params, timeout=20)
            if r.status_code == 200:
                return r.json()
            if r.status_code == 429:
                wait = 2 ** attempt * 5
                logger.warning("Rate limited — waiting %ds", wait)
                time.sleep(wait)
                continue
            body = r.json()
            err = body.get("error", {})
            logger.warning("API %d: %s", r.status_code, err.get("message", r.text[:120]))
            if attempt < MAX_RETRIES - 1:
                time.sleep(2 ** attempt)
        except requests.RequestException as exc:
            logger.warning("Request error (attempt %d/%d): %s", attempt + 1, MAX_RETRIES, exc)
            if attempt < MAX_RETRIES - 1:
                time.sleep(2 ** attempt)
    return {"data": [], "error": "max_retries_exceeded"}


def check_permissions() -> Tuple[bool, List[str]]:
    """
    Validates token permissions. Returns (ok, missing_permissions).
    Required: ads_read + read_insights (via /me/insights access).
    """
    if not ACCESS_TOKEN:
        logger.error("META_ACCESS_TOKEN not set in environment variables.")
        return False, ["META_ACCESS_TOKEN missing"]

    result = api_get(f"{BASE_URL}/me/permissions", {"access_token": ACCESS_TOKEN})
    if "error" in result and "data" not in result:
        return False, [f"Token inválido ou expirado: {result.get('error')}"]

    granted = {p["permission"] for p in result.get("data", []) if p.get("status") == "granted"}
    required = {"ads_read", "ads_management", "business_management"}
    missing = [p for p in required if p not in granted]

    if missing:
        logger.warning("Permissões faltando: %s", missing)
        logger.warning(
            "Gere um novo token em: Meta Business Manager > Configuracoes > "
            "Ferramentas de negócios > Graph API Explorer — escopos: %s",
            ", ".join(required),
        )
    return len(missing) == 0, missing


def fetch_account_info(account_id: str) -> dict:
    return api_get(
        f"{BASE_URL}/{account_id}",
        {
            "fields": "id,name,balance,spend_cap,amount_spent,currency,account_status",
            "access_token": ACCESS_TOKEN,
        },
    )


def fetch_account_insights(account_id: str, date_start: str, date_end: str) -> dict:
    return api_get(
        f"{BASE_URL}/{account_id}/insights",
        {
            "fields": "spend,impressions,clicks,ctr,cpm,cpc,reach,actions,cost_per_action_type",
            "time_range": json.dumps({"since": date_start, "until": date_end}),
            "level": "account",
            "access_token": ACCESS_TOKEN,
        },
    )


def fetch_campaign_insights(account_id: str, date_start: str, date_end: str) -> dict:
    return api_get(
        f"{BASE_URL}/{account_id}/insights",
        {
            "fields": "campaign_name,spend,impressions,clicks,ctr,cpm,cpc,actions",
            "time_range": json.dumps({"since": date_start, "until": date_end}),
            "level": "campaign",
            "access_token": ACCESS_TOKEN,
        },
    )


# ─── Data Processing ───────────────────────────────────────────────────────────

def flt(value, default: float = 0.0) -> float:
    try:
        return float(value) if value else default
    except (ValueError, TypeError):
        return default


def action_val(actions: list, action_type: str) -> float:
    for a in (actions or []):
        if a.get("action_type") == action_type:
            return flt(a.get("value", 0))
    return 0.0


def fmt_brl(v: float) -> str:
    if v == 0:
        return "R$ 0,00"
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def get_status(days: Optional[float]) -> str:
    if days is None:
        return "N/A"
    if days < CRITICAL_DAYS:
        return "CRÍTICO"
    if days < WARNING_DAYS:
        return "ATENÇÃO"
    return "OK"


def campaign_suggestion(camp: dict) -> str:
    spend       = flt(camp.get("spend", 0))
    impressions = flt(camp.get("impressions", 0))
    clicks      = flt(camp.get("clicks", 0))
    ctr         = flt(camp.get("ctr", 0))
    cpm         = flt(camp.get("cpm", 0))
    cpc         = flt(camp.get("cpc", 0))

    if spend == 0:
        return "Sem gasto — verificar se campanha está ativa"
    if clicks == 0 and spend > 50:
        return "Gasto sem cliques — revisar criativo e segmentação"
    if impressions > 5000 and ctr < 0.5:
        return f"CTR crítico ({ctr:.2f}%) — renovar criativos urgentemente"
    if 0 < ctr < CTR_BENCHMARK:
        return f"CTR abaixo do benchmark ({ctr:.2f}% vs {CTR_BENCHMARK}%) — testar novos criativos"
    if cpm > CPM_MAX:
        return f"CPM elevado (R${cpm:.2f}) — refinar segmentação de público"
    if cpc > CPC_MAX:
        return f"CPC alto (R${cpc:.2f}) — otimizar landing page e CTA"
    if ctr > 2.0:
        return f"CTR excelente ({ctr:.2f}%) — escalar orçamento para maximizar resultados"
    return "Performance dentro do esperado — monitorar semanalmente"


def collect_account_data(account: dict, date_start: str, date_end: str) -> dict:
    account_id   = account["id"]
    account_name = account["name"]

    info     = fetch_account_info(account_id)
    ins_raw  = fetch_account_insights(account_id, date_start, date_end)
    camp_raw = fetch_campaign_insights(account_id, date_start, date_end)

    ins  = (ins_raw.get("data") or [{}])[0]
    camps = camp_raw.get("data", [])

    spend_7d       = flt(ins.get("spend", 0))
    avg_daily      = spend_7d / REPORT_DAYS if spend_7d > 0 else 0

    balance        = flt(info.get("balance", 0))
    spend_cap      = flt(info.get("spend_cap", 0))
    amount_spent   = flt(info.get("amount_spent", 0))
    currency       = info.get("currency", "BRL")

    days_remaining: Optional[float] = None
    if balance > 0 and avg_daily > 0:
        days_remaining = balance / avg_daily
    elif spend_cap > 0 and avg_daily > 0:
        remaining = spend_cap - amount_spent
        if remaining > 0:
            days_remaining = remaining / avg_daily

    critical_date = None
    if days_remaining is not None:
        critical_date = (datetime.now() + timedelta(days=days_remaining)).strftime("%d/%m/%Y")

    actions = ins.get("actions", [])
    leads = max(
        action_val(actions, "lead"),
        action_val(actions, "offsite_conversion.fb_pixel_lead"),
        action_val(actions, "onsite_conversion.lead_grouped"),
    )
    messages = action_val(actions, "onsite_conversion.messaging_conversation_started_7d")

    return {
        "id": account_id,
        "name": account_name,
        "balance": balance,
        "spend_cap": spend_cap,
        "amount_spent": amount_spent,
        "currency": currency,
        "spend_7d": spend_7d,
        "avg_daily": avg_daily,
        "days_remaining": days_remaining,
        "critical_date": critical_date,
        "status": get_status(days_remaining),
        "impressions": int(flt(ins.get("impressions", 0))),
        "clicks": int(flt(ins.get("clicks", 0))),
        "ctr": flt(ins.get("ctr", 0)),
        "cpm": flt(ins.get("cpm", 0)),
        "cpc": flt(ins.get("cpc", 0)),
        "reach": int(flt(ins.get("reach", 0))),
        "leads": int(leads + messages),
        "campaigns": [c for c in camps if flt(c.get("spend", 0)) > 0],
        "has_data": spend_7d > 0,
    }


# ─── Excel Builder ─────────────────────────────────────────────────────────────

def _header_row(ws, row: int, headers: List[str]):
    for col, val in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = HEADER_FONT
        c.fill   = BLUE_FILL
        c.alignment = CENTER
        c.border = THIN


def _title_row(ws, row: int, text: str, span: int):
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    c.alignment = CENTER
    ws.row_dimensions[row].height = 32


def _set_col_widths(ws, widths: List[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _status_fill(status: str) -> Optional[PatternFill]:
    return {"CRÍTICO": RED_FILL, "ATENÇÃO": ORANGE_FILL, "OK": GREEN_FILL}.get(status)


def build_tab1(wb: Workbook, data: List[dict], date_start: str, date_end: str):
    ws = wb.active
    ws.title = "Saldos e Projeção"

    _title_row(ws, 1, f"CONTROLE FINANCEIRO META ADS  |  {date_start} → {date_end}", 6)
    _header_row(ws, 2, ["Conta", "Saldo Atual", "Gasto Médio/Dia", "Dias Restantes", "Data Crítica de Recarga", "Status"])
    _set_col_widths(ws, [32, 20, 20, 18, 26, 12])
    ws.row_dimensions[2].height = 36

    status_order = {"CRÍTICO": 0, "ATENÇÃO": 1, "OK": 2, "N/A": 3}
    sorted_data = sorted(data, key=lambda x: (status_order.get(x["status"], 4), x["days_remaining"] or 9999))

    for i, acc in enumerate(sorted_data, 3):
        alt = ALT_FILL if i % 2 == 0 else None
        row_fill = (
            CRIT_LIGHT if acc["status"] == "CRÍTICO" else
            WARN_LIGHT if acc["status"] == "ATENÇÃO" else
            alt
        )
        ws.row_dimensions[i].height = 22

        balance_str = fmt_brl(acc["balance"]) if acc["balance"] > 0 else "Pós-pago / N/A"
        days_str    = f"{acc['days_remaining']:.1f}" if acc["days_remaining"] is not None else "N/A"

        row_vals = [
            (acc["name"],             LEFT),
            (balance_str,             CENTER),
            (fmt_brl(acc["avg_daily"]), CENTER),
            (days_str,                CENTER),
            (acc["critical_date"] or "N/A", CENTER),
            (acc["status"],           CENTER),
        ]
        for col, (val, align) in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = align
            c.border = THIN
            if row_fill:
                c.fill = row_fill

        # Override status cell with colour
        sc = ws.cell(row=i, column=6)
        sf = _status_fill(acc["status"])
        if sf:
            sc.fill = sf
            sc.font = Font(bold=True, color="FFFFFF", size=10)

    ws.freeze_panes = "A3"


def build_tab2(wb: Workbook, data: List[dict], date_start: str, date_end: str):
    ws = wb.create_sheet("Gastos por Campanha")

    _title_row(ws, 1, f"GASTOS POR CAMPANHA — ÚLTIMOS {REPORT_DAYS} DIAS  |  {date_start} → {date_end}", 8)
    _header_row(ws, 2, ["Conta", "Campanha", "Gasto Total 7d", "Média Diária", "CPM", "CPC", "CTR", "Sugestão de Melhoria"])
    _set_col_widths(ws, [28, 50, 16, 16, 12, 12, 10, 62])
    ws.row_dimensions[2].height = 36

    row_idx = 3
    for acc in data:
        for camp in acc["campaigns"]:
            spend = flt(camp.get("spend", 0))
            ctr   = flt(camp.get("ctr", 0))
            ws.row_dimensions[row_idx].height = 36
            alt = ALT_FILL if row_idx % 2 == 0 else None

            row_vals = [
                (acc["name"],                              LEFT),
                (camp.get("campaign_name", "N/A"),         LEFT),
                (fmt_brl(spend),                           CENTER),
                (fmt_brl(spend / REPORT_DAYS),             CENTER),
                (fmt_brl(flt(camp.get("cpm", 0))),         CENTER),
                (fmt_brl(flt(camp.get("cpc", 0))),         CENTER),
                (f"{ctr:.2f}%",                            CENTER),
                (campaign_suggestion(camp),                LEFT),
            ]
            for col, (val, align) in enumerate(row_vals, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.alignment = align
                c.border = THIN
                if alt:
                    c.fill = alt

            # CTR colour coding
            ctr_cell = ws.cell(row=row_idx, column=7)
            if 0 < ctr < 0.5:
                ctr_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                ctr_cell.font = Font(color="B71C1C", bold=True)
            elif ctr > 2.0:
                ctr_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                ctr_cell.font = Font(color="1B5E20", bold=True)

            row_idx += 1

    ws.freeze_panes = "A3"


def build_tab3(wb: Workbook, data: List[dict], date_start: str, date_end: str):
    ws = wb.create_sheet("Resumo Executivo")
    _set_col_widths(ws, [38, 55, 20, 15])

    total_spend       = sum(a["spend_7d"]   for a in data)
    total_impressions = sum(a["impressions"] for a in data)
    total_clicks      = sum(a["clicks"]      for a in data)
    total_leads       = sum(a["leads"]       for a in data)
    active_accounts   = sum(1 for a in data if a["has_data"])
    avg_ctr           = (total_clicks / total_impressions * 100) if total_impressions > 0 else 0
    avg_cpl           = (total_spend / total_leads) if total_leads > 0 else 0

    critical = [a for a in data if a["status"] == "CRÍTICO"]
    warning  = [a for a in data if a["status"] == "ATENÇÃO"]

    all_camps = []
    for acc in data:
        for c in acc["campaigns"]:
            all_camps.append({
                "account": acc["name"],
                "name":    c.get("campaign_name", "N/A"),
                "ctr":     flt(c.get("ctr", 0)),
                "cpc":     flt(c.get("cpc", 0)),
                "cpm":     flt(c.get("cpm", 0)),
                "spend":   flt(c.get("spend", 0)),
                "impressions": flt(c.get("impressions", 0)),
            })
    underperforming = sorted(
        [c for c in all_camps if (0 < c["ctr"] < CTR_BENCHMARK and c["impressions"] > 5000) or c["cpc"] > CPC_MAX],
        key=lambda x: x["ctr"],
    )[:15]

    row = 1

    def section_title(text: str, fill: PatternFill = BLUE_FILL, span: int = 4):
        nonlocal row
        ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = fill
        c.alignment = CENTER
        ws.row_dimensions[row].height = 28
        row += 1

    def data_row(label: str, value: str, value_fill: Optional[PatternFill] = None):
        nonlocal row
        ws.row_dimensions[row].height = 22
        lc = ws.cell(row=row, column=1, value=label)
        lc.font   = Font(size=11)
        lc.alignment = LEFT
        lc.border = THIN
        vc = ws.cell(row=row, column=2, value=value)
        vc.font   = Font(size=11, bold=True)
        vc.alignment = CENTER
        vc.border = THIN
        if value_fill:
            vc.fill = value_fill
            vc.font = Font(size=11, bold=True, color="FFFFFF")
        row += 1

    # Header
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="RESUMO EXECUTIVO — META ADS")
    c.font = Font(bold=True, size=16, color="1877F2")
    c.alignment = CENTER
    ws.row_dimensions[row].height = 38
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1,
                value=f"Período: {date_start} a {date_end}  |  Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = Font(size=10, color="888888")
    c.alignment = CENTER
    row += 2

    # Consolidated metrics
    section_title("MÉTRICAS CONSOLIDADAS (ÚLTIMOS 7 DIAS)")
    data_row("Total Investido",        fmt_brl(total_spend))
    data_row("Contas Ativas",          str(active_accounts))
    data_row("Total de Impressões",    f"{total_impressions:,}".replace(",", "."))
    data_row("Total de Cliques",       f"{total_clicks:,}".replace(",", "."))
    data_row("Total de Leads",         str(total_leads))
    data_row("CTR Médio",              f"{avg_ctr:.2f}%")
    data_row("Custo por Lead Médio",   fmt_brl(avg_cpl))
    data_row("Contas CRÍTICAS",        str(len(critical)), RED_FILL    if critical else None)
    data_row("Contas em ATENÇÃO",      str(len(warning)),  ORANGE_FILL if warning  else None)
    row += 1

    # Risk accounts
    if critical or warning:
        section_title("CONTAS EM RISCO — ORDENADAS POR URGÊNCIA", RED_FILL)
        _header_row(ws, row, ["Conta", "Dias Restantes", "Data Crítica de Recarga", "Status"])
        row += 1
        for acc in critical + warning:
            days_str = f"{acc['days_remaining']:.1f}" if acc["days_remaining"] is not None else "N/A"
            vals = [acc["name"], days_str, acc["critical_date"] or "N/A", acc["status"]]
            aligns = [LEFT, CENTER, CENTER, CENTER]
            for col, (v, a) in enumerate(zip(vals, aligns), 1):
                c = ws.cell(row=row, column=col, value=v)
                c.alignment = a
                c.border = THIN
            sf = _status_fill(acc["status"])
            if sf:
                sc = ws.cell(row=row, column=4)
                sc.fill = sf
                sc.font = Font(bold=True, color="FFFFFF")
            row += 1
        row += 1

    # Underperforming campaigns
    if underperforming:
        section_title("CAMPANHAS COM PERFORMANCE ABAIXO DA MÉDIA", ORANGE_FILL)
        _header_row(ws, row, ["Conta", "Campanha", "CTR", "CPC"])
        row += 1
        for camp in underperforming:
            vals = [camp["account"], camp["name"][:60], f"{camp['ctr']:.2f}%", fmt_brl(camp["cpc"])]
            aligns = [LEFT, LEFT, CENTER, CENTER]
            for col, (v, a) in enumerate(zip(vals, aligns), 1):
                c = ws.cell(row=row, column=col, value=v)
                c.alignment = a
                c.border = THIN
            ctr_c = ws.cell(row=row, column=3)
            ctr_c.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            ctr_c.font = Font(color="B71C1C", bold=True)
            row += 1


# ─── Terminal Summary ──────────────────────────────────────────────────────────

def print_summary(data: List[dict]):
    total_spend = sum(a["spend_7d"] for a in data)
    total_leads = sum(a["leads"]    for a in data)
    critical    = [a for a in data if a["status"] == "CRÍTICO"]
    warning     = [a for a in data if a["status"] == "ATENÇÃO"]

    sep = "=" * 65
    print(f"\n{sep}")
    print("  RESUMO EXECUTIVO — META ADS CONTROL")
    print(sep)
    print(f"  Investimento total (7 dias) : {fmt_brl(total_spend)}")
    print(f"  Total de leads              : {total_leads}")
    print(f"  Contas ativas               : {sum(1 for a in data if a['has_data'])}/{len(data)}")

    if critical:
        print(f"\n  {'⚠ CONTAS CRÍTICAS (< 3 dias)':}")
        for a in critical:
            days_str = f"{a['days_remaining']:.1f}d" if a["days_remaining"] is not None else "N/A"
            print(f"    • {a['name']:<35} {days_str:>8}  →  recarregar até {a['critical_date'] or 'IMEDIATO'}")

    if warning:
        print(f"\n  CONTAS EM ATENÇÃO (< 7 dias):")
        for a in warning:
            days_str = f"{a['days_remaining']:.1f}d" if a["days_remaining"] is not None else "N/A"
            print(f"    • {a['name']:<35} {days_str:>8}  →  recarregar até {a['critical_date'] or 'N/A'}")

    if not critical and not warning:
        print("\n  Todas as contas com saldo OK.")

    print(sep)


# ─── Main ───────────────────────────────────────────────────────────────────────

def main(date_start: Optional[str] = None, date_end: Optional[str] = None):
    if not ACCESS_TOKEN:
        print("\nERRO: Variável META_ACCESS_TOKEN não configurada.")
        print("  Crie um arquivo .env com: META_ACCESS_TOKEN=seu_token_aqui")
        print("  Ou execute: set META_ACCESS_TOKEN=seu_token_aqui  (Windows cmd)")
        print("  Ou execute: $env:META_ACCESS_TOKEN='...'           (PowerShell)\n")
        sys.exit(1)

    # Permission check
    print("\nVerificando permissões do token...")
    ok, missing = check_permissions()
    if not ok:
        print(f"AVISO: Permissões faltando: {missing}")
        print("  Gere novo token em: Meta Business Manager > Ferramentas > Graph API Explorer")
        print("  Escopos necessários: ads_read, ads_management, business_management")
        # Continue anyway — some endpoints may still work with partial permissions

    # Date range
    if not date_start:
        today       = datetime.now()
        last_monday = today - timedelta(days=today.weekday() + 7)
        date_start  = last_monday.strftime("%Y-%m-%d")
        date_end    = (last_monday + timedelta(days=6)).strftime("%Y-%m-%d")

    print(f"\n{'='*65}")
    print(f"  META ADS CONTROL — {date_start} a {date_end}")
    print(f"  Coletando dados de {len(ACCOUNTS)} contas...")
    print(f"{'='*65}")

    accounts_data = []
    for i, account in enumerate(ACCOUNTS):
        print(f"  [{i+1:02d}/{len(ACCOUNTS)}] {account['name']:<40}", end=" ", flush=True)
        acc_data = collect_account_data(account, date_start, date_end)
        accounts_data.append(acc_data)
        status_str = (
            f"R${acc_data['spend_7d']:.0f} | {acc_data['days_remaining']:.1f}d [{acc_data['status']}]"
            if acc_data["days_remaining"] is not None
            else f"R${acc_data['spend_7d']:.0f} [{acc_data['status']}]"
        )
        print(f"→ {status_str}")

    # Build Excel
    reports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
    os.makedirs(reports_dir, exist_ok=True)
    filename    = f"meta_ads_{date_start}.xlsx"
    output_path = os.path.join(reports_dir, filename)

    wb = Workbook()
    build_tab1(wb, accounts_data, date_start, date_end)
    build_tab2(wb, accounts_data, date_start, date_end)
    build_tab3(wb, accounts_data, date_start, date_end)
    wb.save(output_path)

    print_summary(accounts_data)
    print(f"\n  Relatório Excel salvo em:\n  {output_path}\n")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) == 3:
        main(sys.argv[1], sys.argv[2])
    else:
        main()
